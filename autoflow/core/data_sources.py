"""Lecture/écriture de sources de données tabulaires (CSV, Excel) et d'itérables.

Fonctions **pures et testables** (aucune dépendance d'écran). ``openpyxl`` est
importé **paresseusement** : son absence n'empêche pas le CSV ni les listes de
fonctionner, et lève un message clair seulement si un fichier Excel est demandé.

Conçu pour alimenter :
- la boucle **« pour chaque »** (``iter_items``),
- les actions **Lire / Écrire un tableau** (``read_rows`` / ``write_rows``).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


def _is_excel(path: str | Path) -> bool:
    return str(path).lower().endswith((".xlsx", ".xlsm"))


def read_rows(path: str | Path, has_header: bool = True) -> list[dict[str, Any]]:
    """Lit un fichier CSV/Excel en une liste de dictionnaires (par en-tête).

    Si ``has_header`` est faux, les colonnes sont nommées ``col1``, ``col2``…
    Lève :class:`FileNotFoundError` si le fichier est absent, ou
    :class:`RuntimeError` si Excel est demandé sans ``openpyxl``.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")
    rows = _read_excel(path) if _is_excel(path) else _read_csv(path)
    return _as_dicts(rows, has_header)


def _read_csv(path: Path) -> list[list[Any]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return [list(r) for r in csv.reader(fh)]


def _read_excel(path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dépend de l'environnement
        raise RuntimeError(
            "La lecture de fichiers Excel nécessite « openpyxl ».") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    rows = [["" if c is None else c for c in row]
            for row in sheet.iter_rows(values_only=True)]
    wb.close()
    return rows


def _as_dicts(rows: list[list[Any]], has_header: bool) -> list[dict[str, Any]]:
    if not rows:
        return []
    if has_header:
        header = [str(c) for c in rows[0]]
        body = rows[1:]
    else:
        header = [f"col{i + 1}" for i in range(len(rows[0]))]
        body = rows
    result = []
    for row in body:
        item = {header[i]: row[i] if i < len(row) else ""
                for i in range(len(header))}
        result.append(item)
    return result


def write_rows(path: str | Path, rows: list[dict[str, Any]],
               append: bool = False) -> Path:
    """Écrit (ou ajoute) des lignes (dictionnaires) dans un CSV/Excel.

    L'en-tête est déduit des clés de la première ligne. En mode ``append`` sur un
    fichier existant, l'en-tête n'est pas réécrit.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = list(rows or [])
    if _is_excel(path):
        return _write_excel(path, rows, append)
    return _write_csv(path, rows, append)


def _write_csv(path: Path, rows: list[dict[str, Any]], append: bool) -> Path:
    existed = path.exists()
    header = list(rows[0].keys()) if rows else []
    mode = "a" if append and existed else "w"
    with path.open(mode, encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        if header and not (append and existed):
            writer.writerow(header)
        for row in rows:
            writer.writerow([row.get(k, "") for k in header])
    return path


def _write_excel(path: Path, rows: list[dict[str, Any]], append: bool) -> Path:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "L'écriture de fichiers Excel nécessite « openpyxl ».") from exc
    header = list(rows[0].keys()) if rows else []
    if append and path.exists():
        wb = load_workbook(path)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        if header:
            sheet.append(header)
    for row in rows:
        sheet.append([row.get(k, "") for k in header])
    wb.save(path)
    wb.close()
    return path


def iter_items(source: str, value: Any, has_header: bool = True,
               pattern: str = "*") -> list[Any]:
    """Renvoie la liste des éléments à parcourir selon le type de source.

    ``source`` ∈ ``{"list", "variable", "table", "folder", "lines"}`` :

    - ``list`` / ``variable`` : ``value`` est déjà une liste (ou une chaîne
      séparée par des virgules / sauts de ligne).
    - ``table`` : ``value`` est un chemin CSV/Excel → liste de dictionnaires.
    - ``folder`` : ``value`` est un dossier → liste de chemins (filtrés par
      ``pattern``).
    - ``lines`` : ``value`` est un chemin de fichier texte → liste de lignes.
    """
    if source in ("list", "variable"):
        return _coerce_list(value)
    if source == "table":
        return read_rows(value, has_header=has_header)
    if source == "folder":
        folder = Path(str(value))
        if not folder.is_dir():
            return []
        return [str(p) for p in sorted(folder.glob(pattern)) if p.is_file()]
    if source == "lines":
        path = Path(str(value))
        if not path.exists():
            return []
        return path.read_text(encoding="utf-8").splitlines()
    raise ValueError(f"Source de données inconnue : {source!r}")


def _coerce_list(value: Any) -> list[Any]:
    """Normalise une valeur en liste (liste telle quelle, ou chaîne découpée)."""
    if isinstance(value, list):
        return list(value)
    if isinstance(value, tuple):
        return list(value)
    if value is None or value == "":
        return []
    text = str(value)
    if "\n" in text:
        return [line for line in text.splitlines() if line.strip()]
    return [part.strip() for part in text.split(",") if part.strip()]
